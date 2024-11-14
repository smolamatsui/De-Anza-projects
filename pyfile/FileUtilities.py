import os
import zipfile
from pathlib import Path
import shutil
import os
import json
from openpyxl import load_workbook


class Product:
    def __init__(self, name, price):
        self.name = name
        self.price = price
    def __str__(self):
        return self.name + " " + str(self.price)
    def to_bytes(self):
        import pickle
        return pickle.dumps(self)  

text = "A greenhouse gas is a gas that absorbs and emits radiant energy within the thermal infrared range, causing the greenhouse effect. The primary greenhouse gases in Earth's atmosphere are water vapor (H2O), carbon dioxide (CO2), methane (CH4), nitrous oxide (N2O), and ozone. Without greenhouse gases, the average temperature of Earth's surface would be about -18 °C (0 °F), rather than the present average of 15 °C (59 °F).The atmospheres of Venus, Mars and Titan also contain greenhouse gases."

def environment():
    print(len(list(os.environ)), "total environment variables")
    print(list(os.environ))
    print(os.environ.get('OS'))
    print(os.environ.get('NUMBER_OF_PROCESSORS'))
    print(os.environ.get('USERNAME'))
    print(os.environ.get('PROGRAMFILES'))
    print(os.environ.get('PROGRAMFILES(X86)'))

def currentDir():
    return os.getcwd()

def makeDir(nudir):
    return os.mkdir(nudir)

def changeDir(nudir):
    os.chdir(nudir)
    
def removeFile(fname):
    os.remove(fname)
    
def removeDir(dirname):
    os.rmdir(dirname)

def fileList():
    return os.listdir()

def utilUnzip(zfile):  
    with zipfile.ZipFile(zfile,"r") as zf:
        zf.extractall()

def createFile(string):
    fdesc = "temp.txt"
    with open(fdesc, 'w') as file:
        file.write(string)
    return fdesc 

# read created file
def readFile(fname):
    file = open(fname, 'r')
    text = file.read()
    file.close()
    return text

def shCopy(source, destination):
    dest = shutil.copy(source, destination)
    return dest

    
def excelTest():
    workbook = load_workbook(filename="Superstore.xlsx")
    print(f"Worksheet names: {workbook.sheetnames}")
    sheet = workbook.active
    print(sheet)
    print(f"The title of the Worksheet is: {sheet.title}")
    print(f'The value of A2 is {sheet["A2"].value}')
    print(f'The value of A3 is {sheet["A3"].value}')
    cell = sheet['B3']
    print(f'The variable "cell" is {cell.value}')
    workbook.close()
    
def OS_Test():
    print("Current Directory",currentDir())
    makeDir("Sandbox")
    changeDir("Sandbox")
    print("Current Directory",currentDir())
    createFile(text)
    print("File contents",readFile("temp.txt"))
    shCopy("temp.txt","copy1.txt")
    shCopy("temp.txt","copy2.txt")
    shCopy("temp.txt","copy3.txt")
    try:
        shCopy("../Utility.zip", "copy.zip")
    except FileNotFoundError:
        print("Utility.zip not found.")
    print("File List",fileList())
    utilUnzip("copy.zip")
    removeFile("temp.txt")
    filelist = fileList()
    [removeFile(file) for file in filelist]
    print("After files removed: ",fileList())
    changeDir("..")
    removeDir("Sandbox")
 
    
import sqlite3

def databaseTest(product):            
    database = "Utility.db"        
    # Connect to SQLite database
    connection = sqlite3.connect(database)
    cursor = connection.cursor()
    
    # Create table if it doesn't exist
    table_query = '''CREATE TABLE IF NOT EXISTS Blobs (
                        id INTEGER PRIMARY KEY,
                        json TEXT NOT NULL,
                        pickle BLOB NOT NULL UNIQUE)'''
    cursor.execute(table_query)
    
    # Insert data into the table
    insert_query = """INSERT INTO Blobs (id, json, pickle) VALUES (?, ?, ?)"""
    data_t = (1, json.dumps(product.__dict__), product.to_bytes())
    cursor.execute(insert_query, data_t)
    
    # Commit the transaction and close the connection
    connection.commit()
    connection.close()
    
    # Optionally delete the database file to clean up
    os.remove(database)


product = Product('artichoke',0.99)
environment()   
OS_Test()
excelTest()
databaseTest(product)
    
